"""
Модуль работы со справочником сотрудников
Хранение в JSON, управление, экспорт
"""

import json
import os
from config import DOCS_DIR


# Путь к файлу со справочником (внутри docs/)
EMPLOYEES_FILE = os.path.join(DOCS_DIR, "employees.json")


def load_employees():
    """
    Загрузить список сотрудников из JSON

    Returns:
        list[dict]: Список сотрудников
    """
    if not os.path.exists(EMPLOYEES_FILE):
        return []

    try:
        with open(EMPLOYEES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Ошибка загрузки сотрудников: {e}")
        return []


def save_employees(employees):
    """
    Сохранить список сотрудников в JSON

    Args:
        employees: Список сотрудников

    Returns:
        bool: True если успешно
    """
    try:
        with open(EMPLOYEES_FILE, 'w', encoding='utf-8') as f:
            json.dump(employees, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"Ошибка сохранения сотрудников: {e}")
        return False


def get_next_id(employees):
    """
    Получить следующий ID для нового сотрудника

    Args:
        employees: Список сотрудников

    Returns:
        int: Следующий свободный ID
    """
    if not employees:
        return 1

    max_id = max(emp.get('id', 0) for emp in employees)
    return max_id + 1


def add_employee(fio, position, department, email):
    """
    Добавить нового сотрудника

    Args:
        fio: ФИО
        position: Должность
        department: Подразделение (ФБП или НПФ)
        email: Email

    Returns:
        bool: True если успешно
    """
    employees = load_employees()

    new_employee = {
        "id": get_next_id(employees),
        "fio": fio.strip(),
        "position": position.strip(),
        "department": department,
        "email": email.strip()
    }

    employees.append(new_employee)
    return save_employees(employees)


def update_employee(employee_id, fio, position, department, email):
    """
    Обновить данные сотрудника

    Args:
        employee_id: ID сотрудника
        fio: ФИО
        position: Должность
        department: Подразделение
        email: Email

    Returns:
        bool: True если успешно
    """
    employees = load_employees()

    for emp in employees:
        if emp['id'] == employee_id:
            emp['fio'] = fio.strip()
            emp['position'] = position.strip()
            emp['department'] = department
            emp['email'] = email.strip()
            return save_employees(employees)

    return False


def delete_employee(employee_id):
    """
    Удалить сотрудника

    Args:
        employee_id: ID сотрудника

    Returns:
        bool: True если успешно
    """
    employees = load_employees()
    employees = [emp for emp in employees if emp['id'] != employee_id]
    return save_employees(employees)


def export_employees_to_excel(output_path):
    """
    Экспорт списка сотрудников в Excel

    Args:
        output_path: Путь для сохранения

    Returns:
        bool: True если успешно
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime

        employees = load_employees()

        if not employees:
            return False

        # Создаём книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Сотрудники"

        # Стили
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')

        title_font = Font(name='Arial', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')

        normal_font = Font(name='Arial', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовок
        ws.merge_cells('A1:E1')
        ws['A1'] = 'СПРАВОЧНИК СОТРУДНИКОВ'
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment

        ws.merge_cells('A2:E2')
        ws['A2'] = f'Дата экспорта: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}'
        ws['A2'].font = Font(name='Arial', size=10)
        ws['A2'].alignment = Alignment(horizontal='center')

        # Заголовки колонок
        headers = ['№', 'ФИО', 'Должность', 'Подразделение', 'Email']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Ширина колонок
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30

        # Данные
        for idx, emp in enumerate(employees, start=1):
            row = idx + 4

            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = emp['fio']
            ws.cell(row=row, column=3).value = emp['position']
            ws.cell(row=row, column=4).value = emp['department']
            ws.cell(row=row, column=5).value = emp['email']

            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.font = normal_font
                cell.border = border
                if col == 1 or col == 4:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # Итого
        row = len(employees) + 6
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = f'Всего сотрудников: {len(employees)}'
        ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        # Сохраняем
        wb.save(output_path)
        return True

    except ImportError:
        print("Ошибка: библиотека openpyxl не установлена")
        return False
    except Exception as e:
        print(f"Ошибка экспорта: {e}")
        return False


def create_familiarization_sheet(document, selected_employees, output_path):
    """
    Создать лист ознакомления с документом для выбранных сотрудников

    Args:
        document: Объект документа (Document)
        selected_employees: Список выбранных сотрудников (list[dict])
        output_path: Путь для сохранения Excel файла

    Returns:
        bool: True если успешно
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        from logic import build_filename

        if not selected_employees:
            return False

        # Создаём книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Лист ознакомления"

        # Стили
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')

        title_font = Font(name='Arial', size=16, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')

        subtitle_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовок
        ws.merge_cells('A1:F1')
        ws['A1'] = 'ЛИСТ ОЗНАКОМЛЕНИЯ С ДОКУМЕНТОМ СМК'
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment

        # Информация о документе
        if document.is_valid:
            doc_name = build_filename(document.typ, document.kod, document.version, document.year, document.title)
        else:
            doc_name = document.filename

        ws.merge_cells('A3:F3')
        ws['A3'] = f'Документ: {doc_name}'
        ws['A3'].font = subtitle_font
        ws['A3'].alignment = Alignment(horizontal='left')

        if hasattr(document, 'category') and document.category:
            ws.merge_cells('A4:F4')
            ws['A4'] = f'Категория: {document.category}'
            ws['A4'].font = Font(name='Arial', size=11)
            ws['A4'].alignment = Alignment(horizontal='left')

        ws.merge_cells('A5:F5')
        ws['A5'] = f'Дата формирования листа: {datetime.now().strftime("%d.%m.%Y")}'
        ws['A5'].font = Font(name='Arial', size=11)
        ws['A5'].alignment = Alignment(horizontal='left')

        # Заголовки таблицы
        headers = ['№', 'ФИО', 'Должность', 'Подразделение', 'Дата ознакомления', 'Подпись']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=7, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Ширина колонок
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

        # Данные сотрудников
        for idx, emp in enumerate(selected_employees, start=1):
            row = idx + 7

            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = emp['fio']
            ws.cell(row=row, column=3).value = emp['position']
            ws.cell(row=row, column=4).value = emp['department']
            ws.cell(row=row, column=5).value = ""  # Пустая ячейка для даты
            ws.cell(row=row, column=6).value = ""  # Пустая ячейка для подписи

            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.font = normal_font
                cell.border = border
                if col == 1 or col == 4:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # Итого
        row = len(selected_employees) + 9
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f'Всего для ознакомления: {len(selected_employees)} человек(а)'
        ws[f'A{row}'].font = Font(name='Arial', size=11, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        # Сохраняем
        wb.save(output_path)
        return True

    except ImportError:
        print("Ошибка: библиотека openpyxl не установлена")
        return False
    except Exception as e:
        print(f"Ошибка создания листа ознакомления: {e}")
        return False
