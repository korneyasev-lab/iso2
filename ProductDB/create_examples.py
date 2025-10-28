"""
Скрипт для создания примеров файлов Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

# ============================================
# СОЗДАТЬ ПРИМЕР ФАЙЛА ЦЕН (Excel)
# ============================================

def create_prices_example():
    """Создать пример файла цен в Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Цены"

    # Заголовки
    headers = ['ID комплектующего', 'Цена', 'Валюта', 'Дата начала действия', 'Примечания']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Данные
    data = [
        [1, 15.50, 'RUB', '2025-01-01', 'Бетон M20'],
        [2, 18.00, 'RUB', '2025-01-01', 'Бетон M25'],
        [3, 250.00, 'RUB', '2025-01-01', 'Закладная деталь ZD-100'],
        [4, 5.00, 'RUB', '2025-01-01', 'Болт М12'],
        [5, 120.00, 'RUB', '2025-01-01', 'Картон асбестовый'],
    ]

    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = Font(name="Arial", size=10)

    # Настроить ширину колонок
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 35

    wb.save('examples/prices_example.xlsx')
    print("✅ Создан: examples/prices_example.xlsx")


# ============================================
# СОЗДАТЬ ШАБЛОН ППМ
# ============================================

def create_ppm_template():
    """Создать шаблон ППМ для директора"""
    wb = Workbook()
    ws = wb.active
    ws.title = "План"

    # Заголовок
    ws.merge_cells('A1:I1')
    cell = ws['A1']
    cell.value = "ПЛАН ПРОИЗВОДСТВА НА МЕСЯЦ"
    cell.font = Font(name="Arial", size=16, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Информация о плане
    ws['A2'] = "Месяц:"
    ws['A2'].font = Font(name="Arial", size=11, bold=True)
    ws['B2'] = ""  # Заполняется программно

    ws['C2'] = "Год:"
    ws['C2'].font = Font(name="Arial", size=11, bold=True)
    ws['D2'] = ""  # Заполняется программно

    ws['E2'] = "Версия:"
    ws['E2'].font = Font(name="Arial", size=11, bold=True)
    ws['F2'] = ""  # Заполняется программно

    # Заголовки таблицы
    headers = ['№', 'Заказчик', 'Чертеж', 'Вариант', 'Было (шт)', 'Стало (шт)',
               'Срок был', 'Срок стал', 'Примечания']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Примеры данных (строка 5 и далее заполняются программно)
    ws['A5'] = 1
    ws['B5'] = "Пример Заказчик"
    ws['C5'] = "PB 35.126"
    ws['D5'] = "R2-B15-НТ"
    ws['E5'] = 0
    ws['F5'] = 100
    ws['G5'] = ""
    ws['H5'] = "2025-11-30"
    ws['I5'] = "Новый заказ"

    # Настроить ширину колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30

    # Подпись
    ws['A30'] = "Утверждаю:"
    ws['A30'].font = Font(name="Arial", size=11, bold=True)
    ws['A31'] = "Директор"
    ws['E31'] = "________________"
    ws['G31'] = "«___» __________ 20__ г."

    wb.save('templates/PPM_template.xlsx')
    print("✅ Создан: templates/PPM_template.xlsx")


# ============================================
# СОЗДАТЬ КОНФИГУРАЦИЮ ППМ
# ============================================

def create_ppm_config():
    """Создать конфигурацию для заполнения ППМ"""
    import json

    config = {
        "month_cell": "B2",
        "year_cell": "D2",
        "version_cell": "F2",
        "items_start_row": 5,
        "columns": {
            "number": "A",
            "client": "B",
            "drawing": "C",
            "variant": "D",
            "quantity_was": "E",
            "quantity_now": "F",
            "deadline_was": "G",
            "deadline_now": "H",
            "notes": "I"
        }
    }

    with open('templates/PPM_config.json', 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=4)

    print("✅ Создан: templates/PPM_config.json")


# ============================================
# ЗАПУСК
# ============================================

if __name__ == "__main__":
    print("🔧 Создание примеров файлов...")
    create_prices_example()
    create_ppm_template()
    create_ppm_config()
    print("\n✅ Все файлы созданы!")
