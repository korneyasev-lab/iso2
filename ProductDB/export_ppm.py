"""
ProductDB - PPM Export Module
Модуль экспорта отчётов ППМ в Excel
"""

import os
from datetime import datetime
from typing import Dict, List, Optional
import json

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  ВНИМАНИЕ: Библиотека openpyxl не установлена")
    print("   Установите: pip install openpyxl")

from database import Database
from logic_ppm import calculate_material_requirements, calculate_costs


# ============================================
# КОНСТАНТЫ
# ============================================

# Стили для Excel
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SUBHEADER_FONT = Font(name="Arial", size=10, bold=True)
NORMAL_FONT = Font(name="Arial", size=10)
TOTAL_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
TOTAL_FONT = Font(name="Arial", size=11, bold=True)

BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# ============================================
# ОТЧЕТ ДЛЯ СНАБЖЕНЦА
# ============================================

def export_supply_report(plan_id: int, db: Database, filepath: str = None,
                        include_prices: bool = True) -> str:
    """
    Экспортировать отчет для снабженца

    Формат отчета:
    - Заголовок с информацией о плане
    - Таблица материалов по типам
    - Колонки: №, Наименование, Единица, Количество, [Цена, Сумма]
    - Итоговая строка

    Args:
        plan_id: ID плана производства
        db: Объект подключения к БД
        filepath: Путь для сохранения (по умолчанию генерируется автоматически)
        include_prices: Включать ли цены в отчет

    Returns:
        Путь к созданному файлу
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("Библиотека openpyxl не установлена")

    # Рассчитать потребность
    requirements = calculate_material_requirements(plan_id, db)
    plan_info = requirements['plan_info']

    # Рассчитать затраты (если нужны цены)
    costs_data = None
    if include_prices:
        costs_data = calculate_costs(requirements, db)

    # Сгенерировать имя файла
    if filepath is None:
        month = plan_info['month']
        year = plan_info['year']
        version = plan_info['version']
        os.makedirs('reports/supply', exist_ok=True)
        filepath = f"reports/supply/Потребность_материалы_{year}-{month:02d}_v{version}.xlsx"

    # Создать книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Потребность в материалах"

    # Настроить ширину колонок
    ws.column_dimensions['A'].width = 5   # №
    ws.column_dimensions['B'].width = 40  # Наименование
    ws.column_dimensions['C'].width = 12  # Единица
    ws.column_dimensions['D'].width = 15  # Количество
    if include_prices:
        ws.column_dimensions['E'].width = 15  # Цена
        ws.column_dimensions['F'].width = 18  # Сумма

    # Заголовок отчета
    row = 1
    ws.merge_cells(f'A{row}:{"F" if include_prices else "D"}{row}')
    cell = ws[f'A{row}']
    cell.value = f"ПОТРЕБНОСТЬ В МАТЕРИАЛАХ"
    cell.font = Font(name="Arial", size=14, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    # Информация о плане
    ws.merge_cells(f'A{row}:{"F" if include_prices else "D"}{row}')
    cell = ws[f'A{row}']
    month_names = ['', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
                   'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
    cell.value = f"План: {month_names[plan_info['month']]} {plan_info['year']} (версия {plan_info['version']})"
    cell.font = Font(name="Arial", size=11)
    cell.alignment = Alignment(horizontal='center')
    row += 1

    ws.merge_cells(f'A{row}:{"F" if include_prices else "D"}{row}')
    cell = ws[f'A{row}']
    cell.value = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    cell.font = Font(name="Arial", size=10, italic=True)
    cell.alignment = Alignment(horizontal='center')
    row += 1

    # Пустая строка
    row += 1

    # Заголовок таблицы
    headers = ['№', 'Наименование', 'Ед.', 'Количество']
    if include_prices:
        headers.extend(['Цена, руб.', 'Сумма, руб.'])

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_THIN

    row += 1

    # Данные таблицы
    item_num = 1
    total_cost = 0.0

    # Группировать по типу комплектующих
    from logic_ppm import group_by_component_type
    grouped = group_by_component_type(requirements['requirements'])

    for comp_type, items in sorted(grouped.items()):
        # Подзаголовок типа
        ws.merge_cells(f'A{row}:{"F" if include_prices else "D"}{row}')
        cell = ws[f'A{row}']
        cell.value = comp_type.upper()
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = BORDER_THIN
        row += 1

        # Материалы этого типа
        for item in items:
            # Номер
            cell = ws.cell(row=row, column=1)
            cell.value = item_num
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = BORDER_THIN

            # Наименование
            cell = ws.cell(row=row, column=2)
            name = item['component_name']
            if item['concrete_brand']:
                name += f" ({item['concrete_brand']})"
            cell.value = name
            cell.font = NORMAL_FONT
            cell.border = BORDER_THIN

            # Единица измерения
            cell = ws.cell(row=row, column=3)
            cell.value = item['unit']
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = BORDER_THIN

            # Количество
            cell = ws.cell(row=row, column=4)
            cell.value = round(item['quantity'], 2)
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal='right')
            cell.border = BORDER_THIN
            cell.number_format = '0.00'

            # Цена и сумма (если нужно)
            if include_prices and costs_data:
                # Найти соответствующий элемент в costs_data
                cost_item = next(
                    (c for c in costs_data['requirements_with_prices']
                     if c['component_id'] == item['component_id']),
                    None
                )

                if cost_item:
                    # Цена
                    cell = ws.cell(row=row, column=5)
                    cell.value = round(cost_item['price'], 2)
                    cell.font = NORMAL_FONT
                    cell.alignment = Alignment(horizontal='right')
                    cell.border = BORDER_THIN
                    cell.number_format = '0.00'

                    # Сумма
                    cell = ws.cell(row=row, column=6)
                    cell.value = round(cost_item['total_cost'], 2)
                    cell.font = NORMAL_FONT
                    cell.alignment = Alignment(horizontal='right')
                    cell.border = BORDER_THIN
                    cell.number_format = '0.00'

                    total_cost += cost_item['total_cost']

            item_num += 1
            row += 1

    # Итоговая строка
    if include_prices:
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "ИТОГО:"
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = BORDER_THIN

        cell = ws.cell(row=row, column=6)
        cell.value = round(total_cost, 2)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.alignment = Alignment(horizontal='right')
        cell.border = BORDER_THIN
        cell.number_format = '0.00'
        row += 1

    # Бетон по маркам (дополнительная информация)
    if requirements['concrete_by_brand']:
        row += 2
        ws.merge_cells(f'A{row}:{"F" if include_prices else "D"}{row}')
        cell = ws[f'A{row}']
        cell.value = "БЕТОН ПО МАРКАМ"
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='left')
        row += 1

        for brand, weight in sorted(requirements['concrete_by_brand'].items()):
            cell = ws.cell(row=row, column=2)
            cell.value = f"{brand}:"
            cell.font = NORMAL_FONT

            cell = ws.cell(row=row, column=3)
            cell.value = round(weight, 2)
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '0.00'

            cell = ws.cell(row=row, column=4)
            cell.value = "кг"
            cell.font = NORMAL_FONT
            row += 1

    # Сохранить файл
    wb.save(filepath)
    return filepath


# ============================================
# ППМ ДЛЯ ДИРЕКТОРА
# ============================================

def export_ppm_for_director(plan_id: int, db: Database,
                           template_path: str = "templates/PPM_template.xlsx",
                           config_path: str = "templates/PPM_config.json",
                           output_path: str = None) -> str:
    """
    Экспортировать ППМ для директора на основе шаблона

    Использует Excel шаблон и JSON конфигурацию для заполнения ячеек

    Args:
        plan_id: ID плана производства
        db: Объект подключения к БД
        template_path: Путь к шаблону Excel
        config_path: Путь к конфигурации JSON
        output_path: Путь для сохранения (по умолчанию генерируется автоматически)

    Returns:
        Путь к созданному файлу
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("Библиотека openpyxl не установлена")

    # Получить данные плана
    plan = db.get_production_plan(plan_id)
    if not plan:
        raise ValueError(f"План с ID {plan_id} не найден")

    items = db.get_plan_items(plan_id)

    # Проверить наличие шаблона
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")

    # Загрузить конфигурацию
    config = {}
    if os.path.exists(config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)

    # Загрузить шаблон
    wb = load_workbook(template_path)
    ws = wb.active

    # Заполнить данные согласно конфигурации
    # Пример конфигурации:
    # {
    #   "month_cell": "B2",
    #   "year_cell": "D2",
    #   "version_cell": "F2",
    #   "items_start_row": 5,
    #   "columns": {
    #     "client": "A",
    #     "drawing": "B",
    #     "variant": "C",
    #     "quantity": "D",
    #     "deadline": "E"
    #   }
    # }

    # Заполнить заголовок
    if 'month_cell' in config:
        ws[config['month_cell']] = plan['month']
    if 'year_cell' in config:
        ws[config['year_cell']] = plan['year']
    if 'version_cell' in config:
        ws[config['version_cell']] = plan['version']

    # Заполнить позиции
    if 'items_start_row' in config and 'columns' in config:
        row = config['items_start_row']
        cols = config['columns']

        for i, item in enumerate(items, 1):
            if 'number' in cols:
                ws[f"{cols['number']}{row}"] = i
            if 'client' in cols:
                ws[f"{cols['client']}{row}"] = item['client_name']
            if 'drawing' in cols:
                ws[f"{cols['drawing']}{row}"] = item['drawing_number']
            if 'variant' in cols:
                ws[f"{cols['variant']}{row}"] = item['variant']
            if 'quantity_was' in cols:
                ws[f"{cols['quantity_was']}{row}"] = item['quantity_was']
            if 'quantity_now' in cols:
                ws[f"{cols['quantity_now']}{row}"] = item['quantity_now']
            if 'deadline_was' in cols:
                ws[f"{cols['deadline_was']}{row}"] = item['deadline_was']
            if 'deadline_now' in cols:
                ws[f"{cols['deadline_now']}{row}"] = item['deadline_now']
            if 'notes' in cols:
                ws[f"{cols['notes']}{row}"] = item['notes'] or ''

            row += 1

    # Сгенерировать имя файла
    if output_path is None:
        month = plan['month']
        year = plan['year']
        version = plan['version']
        os.makedirs('reports/plans', exist_ok=True)
        output_path = f"reports/plans/ППМ_{year}-{month:02d}_v{version}.xlsx"

    # Сохранить
    wb.save(output_path)
    return output_path


# ============================================
# ИМПОРТ ЦЕН
# ============================================

def import_prices_from_excel(filepath: str, db: Database) -> tuple:
    """
    Импортировать цены из Excel файла

    Ожидаемый формат:
    - Первая строка - заголовки (пропускается)
    - Колонки: ID комплектующего, Цена, Валюта, Дата начала действия, Примечания

    Args:
        filepath: Путь к Excel файлу
        db: Объект подключения к БД

    Returns:
        (количество добавленных цен, список ошибок)
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("Библиотека openpyxl не установлена")

    wb = load_workbook(filepath)
    ws = wb.active

    prices_data = []
    errors = []

    # Пропустить заголовок
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row[0]:  # Пустая строка
                continue

            price_dict = {
                'component_id': int(row[0]),
                'price': float(row[1]),
                'currency': row[2] if len(row) > 2 and row[2] else 'RUB',
                'valid_from': row[3] if len(row) > 3 and row[3] else None,
                'notes': row[4] if len(row) > 4 and row[4] else None
            }

            # Конвертировать дату если нужно
            if isinstance(price_dict['valid_from'], datetime):
                price_dict['valid_from'] = price_dict['valid_from'].date().isoformat()

            prices_data.append(price_dict)
        except Exception as e:
            errors.append(f"Строка {row_num}: {str(e)}")

    # Импортировать
    count, import_errors = db.import_prices_from_dict(prices_data)
    errors.extend(import_errors)

    return count, errors


def import_prices_from_json(filepath: str, db: Database) -> tuple:
    """
    Импортировать цены из JSON файла

    Ожидаемый формат:
    [
        {
            "component_id": 1,
            "price": 15.50,
            "currency": "RUB",
            "valid_from": "2025-01-01",
            "notes": "Цена с начала года"
        },
        ...
    ]

    Args:
        filepath: Путь к JSON файлу
        db: Объект подключения к БД

    Returns:
        (количество добавленных цен, список ошибок)
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        prices_data = json.load(f)

    return db.import_prices_from_dict(prices_data)


def parse_price_file(filepath: str, db: Database) -> tuple:
    """
    Автоматически определить формат файла и импортировать цены

    Args:
        filepath: Путь к файлу (Excel или JSON)
        db: Объект подключения к БД

    Returns:
        (количество добавленных цен, список ошибок)
    """
    ext = os.path.splitext(filepath)[1].lower()

    if ext in ['.xlsx', '.xls']:
        return import_prices_from_excel(filepath, db)
    elif ext == '.json':
        return import_prices_from_json(filepath, db)
    else:
        raise ValueError(f"Неподдерживаемый формат файла: {ext}")


# ============================================
# ТЕСТИРОВАНИЕ
# ============================================

if __name__ == "__main__":
    print("🔧 Тестирование export_ppm.py...")

    if not OPENPYXL_AVAILABLE:
        print("❌ Установите openpyxl: pip install openpyxl")
        exit(1)

    with Database() as db:
        # Получить тестовый план
        plans = db.get_all_production_plans()
        if not plans:
            print("❌ Нет планов для тестирования")
            exit(1)

        plan = plans[0]
        plan_id = plan['id']

        print(f"\n📋 План ID={plan_id}: {plan['month']}/{plan['year']} v{plan['version']}")

        # Тест 1: Экспорт отчета для снабженца (с ценами)
        print("\n1️⃣ Экспорт отчета для снабженца (с ценами)...")
        try:
            filepath = export_supply_report(plan_id, db, include_prices=True)
            print(f"   ✅ Отчет создан: {filepath}")
            print(f"   📊 Размер файла: {os.path.getsize(filepath)} байт")
        except Exception as e:
            print(f"   ❌ Ошибка: {e}")

        # Тест 2: Экспорт отчета для снабженца (без цен)
        print("\n2️⃣ Экспорт отчета для снабженца (без цен)...")
        try:
            filepath = export_supply_report(
                plan_id, db,
                filepath="reports/supply/Потребность_материалы_test_no_prices.xlsx",
                include_prices=False
            )
            print(f"   ✅ Отчет создан: {filepath}")
            print(f"   📊 Размер файла: {os.path.getsize(filepath)} байт")
        except Exception as e:
            print(f"   ❌ Ошибка: {e}")

        # Тест 3: Экспорт ППМ для директора (если есть шаблон)
        print("\n3️⃣ Экспорт ППМ для директора...")
        if os.path.exists("templates/PPM_template.xlsx"):
            try:
                filepath = export_ppm_for_director(plan_id, db)
                print(f"   ✅ ППМ создан: {filepath}")
            except Exception as e:
                print(f"   ❌ Ошибка: {e}")
        else:
            print("   ⚠️  Шаблон не найден (templates/PPM_template.xlsx)")
            print("   💡 Создайте шаблон вручную для тестирования этой функции")

    print("\n✅ Тестирование завершено!")
