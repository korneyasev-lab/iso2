"""
Модуль работы с реестрами документации
Просмотр, экспорт в CSV и Excel
"""

import os
import csv
from datetime import datetime
from config import (
    CATEGORIES, REGISTRY_ACTUAL_FILES, ACTIVE_CATEGORIES,
    REGISTRIES_CATEGORIES
)
from logic import scan_folder, build_filename, create_registry_for_category


def read_registry_content(category):
    """
    Прочитать содержимое актуального реестра для категории

    Args:
        category: Категория документа

    Returns:
        str: Содержимое реестра или сообщение об отсутствии
    """
    if category not in REGISTRY_ACTUAL_FILES:
        return "Категория не найдена"

    registry_path = REGISTRY_ACTUAL_FILES[category]

    if not os.path.exists(registry_path):
        return f"Реестр для категории '{category}' ещё не создан.\nОпубликуйте первый документ в эту категорию."

    try:
        with open(registry_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content
    except Exception as e:
        return f"Ошибка чтения реестра: {e}"


def get_registry_documents(category):
    """
    Получить список документов из реестра в виде структурированных данных

    Args:
        category: Категория документа

    Returns:
        list[dict]: Список документов [{номер, название}, ...]
    """
    if category not in REGISTRY_ACTUAL_FILES:
        return []

    registry_path = REGISTRY_ACTUAL_FILES[category]

    if not os.path.exists(registry_path):
        return []

    documents = []

    try:
        with open(registry_path, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        # Ищем строки с документами (формат: "1. Название документа")
        for line in lines:
            line = line.strip()
            # Пропускаем заголовки и разделители
            if not line or line.startswith('═') or 'РЕЕСТР' in line or 'Категория:' in line or 'Версия:' in line or 'Дата:' in line or 'ИТОГО:' in line:
                continue

            # Проверяем формат "номер. название"
            if '. ' in line:
                parts = line.split('. ', 1)
                if len(parts) == 2 and parts[0].isdigit():
                    documents.append({
                        'номер': parts[0],
                        'название': parts[1]
                    })

        return documents

    except Exception as e:
        print(f"Ошибка парсинга реестра: {e}")
        return []


def export_registry_to_csv(category, output_path):
    """
    Экспорт реестра в CSV

    Args:
        category: Категория документа
        output_path: Путь для сохранения CSV файла

    Returns:
        bool: True если успешно
    """
    try:
        documents = get_registry_documents(category)

        if not documents:
            return False

        with open(output_path, 'w', encoding='utf-8-sig', newline='') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')

            # Заголовок
            writer.writerow(['Реестр действующих документов СМК'])
            writer.writerow([f'Категория: {category}'])
            writer.writerow([f'Дата экспорта: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}'])
            writer.writerow([])  # Пустая строка
            writer.writerow(['№', 'Название документа'])

            # Документы
            for doc in documents:
                writer.writerow([doc['номер'], doc['название']])

            writer.writerow([])
            writer.writerow([f'ИТОГО: {len(documents)} документов'])

        return True

    except Exception as e:
        print(f"Ошибка экспорта в CSV: {e}")
        return False


def export_registry_to_excel(category, output_path):
    """
    Экспорт реестра в Excel с форматированием

    Args:
        category: Категория документа
        output_path: Путь для сохранения Excel файла

    Returns:
        bool: True если успешно
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        documents = get_registry_documents(category)

        if not documents:
            return False

        # Создаём книгу
        wb = Workbook()
        ws = wb.active
        ws.title = category[:31]  # Excel ограничение на длину имени листа

        # Стили
        header_font = Font(name='Arial', size=14, bold=True)
        header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')

        title_font = Font(name='Arial', size=16, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')

        normal_font = Font(name='Arial', size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовок
        ws.merge_cells('A1:B1')
        ws['A1'] = 'РЕЕСТР ДЕЙСТВУЮЩИХ ДОКУМЕНТОВ СМК'
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment

        ws.merge_cells('A2:B2')
        ws['A2'] = f'Категория: {category}'
        ws['A2'].font = Font(name='Arial', size=12, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:B3')
        ws['A3'] = f'Дата экспорта: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}'
        ws['A3'].font = normal_font
        ws['A3'].alignment = Alignment(horizontal='center')

        # Заголовки колонок
        ws['A5'] = '№'
        ws['B5'] = 'Название документа'

        ws['A5'].font = header_font
        ws['B5'].font = header_font
        ws['A5'].fill = header_fill
        ws['B5'].fill = header_fill
        ws['A5'].alignment = header_alignment
        ws['B5'].alignment = header_alignment
        ws['A5'].border = border
        ws['B5'].border = border

        # Ширина колонок
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 100

        # Данные
        row = 6
        for doc in documents:
            ws[f'A{row}'] = doc['номер']
            ws[f'B{row}'] = doc['название']

            ws[f'A{row}'].font = normal_font
            ws[f'B{row}'].font = normal_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='top')
            ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border

            row += 1

        # Итого
        row += 1
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'] = f'ИТОГО: {len(documents)} действующих документов'
        ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        # Сохраняем
        wb.save(output_path)
        return True

    except ImportError:
        print("Ошибка: библиотека openpyxl не установлена. Установите: pip install openpyxl")
        return False
    except Exception as e:
        print(f"Ошибка экспорта в Excel: {e}")
        return False


def export_all_registries_to_excel(output_path):
    """
    Экспорт всех реестров в один Excel файл (каждая категория = отдельный лист)

    Args:
        output_path: Путь для сохранения Excel файла

    Returns:
        bool: True если успешно
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        wb.remove(wb.active)  # Удаляем дефолтный лист

        # Стили
        header_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')

        title_font = Font(name='Arial', size=16, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')

        normal_font = Font(name='Arial', size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Для каждой категории создаём лист
        for category in CATEGORIES:
            documents = get_registry_documents(category)

            if not documents:
                continue  # Пропускаем пустые реестры

            # Создаём лист
            ws = wb.create_sheet(title=category[:31])

            # Заголовок
            ws.merge_cells('A1:B1')
            ws['A1'] = 'РЕЕСТР ДЕЙСТВУЮЩИХ ДОКУМЕНТОВ СМК'
            ws['A1'].font = title_font
            ws['A1'].alignment = title_alignment

            ws.merge_cells('A2:B2')
            ws['A2'] = f'Категория: {category}'
            ws['A2'].font = Font(name='Arial', size=12, bold=True)
            ws['A2'].alignment = Alignment(horizontal='center')

            ws.merge_cells('A3:B3')
            ws['A3'] = f'Дата экспорта: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}'
            ws['A3'].font = normal_font
            ws['A3'].alignment = Alignment(horizontal='center')

            # Заголовки колонок
            ws['A5'] = '№'
            ws['B5'] = 'Название документа'

            ws['A5'].font = header_font
            ws['B5'].font = header_font
            ws['A5'].fill = header_fill
            ws['B5'].fill = header_fill
            ws['A5'].alignment = header_alignment
            ws['B5'].alignment = header_alignment
            ws['A5'].border = border
            ws['B5'].border = border

            # Ширина колонок
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 100

            # Данные
            row = 6
            for doc in documents:
                ws[f'A{row}'] = doc['номер']
                ws[f'B{row}'] = doc['название']

                ws[f'A{row}'].font = normal_font
                ws[f'B{row}'].font = normal_font
                ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='top')
                ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border

                row += 1

            # Итого
            row += 1
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = f'ИТОГО: {len(documents)} действующих документов'
            ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')

        # Сохраняем
        if len(wb.sheetnames) > 0:
            wb.save(output_path)
            return True
        else:
            return False

    except ImportError:
        print("Ошибка: библиотека openpyxl не установлена. Установите: pip install openpyxl")
        return False
    except Exception as e:
        print(f"Ошибка экспорта всех реестров: {e}")
        return False


def manual_update_registry(category):
    """
    Принудительное обновление реестра для категории
    (без публикации документа)

    Args:
        category: Категория документа

    Returns:
        bool: True если успешно
    """
    try:
        create_registry_for_category(category)
        return True
    except Exception as e:
        print(f"Ошибка обновления реестра: {e}")
        return False
